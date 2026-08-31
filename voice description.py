"""Локальная расшифровка разговора с разделением по собеседникам.

Публичная функция :func:`analyze_call` принимает имя аудиофайла и создаёт
рядом Excel-файл с колонками «Таймлайн», «Собеседник», «Что сказано».

Распознавание выполняется локально через faster-whisper, диаризация — через
sherpa-onnx. Этот модуль не выполняет сетевых запросов и не требует API-ключей.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


SCRIPT_DIR = Path(__file__).resolve().parent
SETTINGS_FILE = SCRIPT_DIR / "voice settings.json"
MODELS_DIR = SCRIPT_DIR / "models"
WHISPER_MODEL_DIR = MODELS_DIR / "faster-whisper-small"
SEGMENTATION_MODEL = MODELS_DIR / "sherpa-onnx-pyannote-segmentation-3-0" / "model.int8.onnx"
EMBEDDING_MODEL = MODELS_DIR / "3dspeaker_speech_eres2net_base_sv_zh-cn_3dspeaker_16k.onnx"
EXPECTED_MODEL_SIZES = {
    WHISPER_MODEL_DIR / "model.bin": 483_546_902,
    SEGMENTATION_MODEL: 1_540_506,
    EMBEDDING_MODEL: 39_593_761,
}

SAMPLE_RATE = 16_000
CYRILLIC_LABELS = tuple("АБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ")
SPEAKER_COLORS = ("E8F1FB", "FCE8E6", "E6F4EA", "FFF4D6", "F3E8FD", "E0F7FA")
SUPPORTED_EXTENSIONS = {".flac", ".m4a", ".mp3", ".mp4", ".mpeg", ".mpga", ".ogg", ".wav", ".webm"}

DEFAULT_SETTINGS: dict[str, Any] = {
    "limits": {
        "max_file_size_mb": 0,
        "max_duration_minutes": 0,
    },
    "recognition": {
        "language": "ru",
        "device": "cpu",
        "compute_type": "int8",
        "cpu_threads": 0,
        "beam_size": 5,
        "temperature": 0.0,
        "condition_on_previous_text": True,
        "vad_filter": True,
        "vad_min_silence_duration_ms": 200,
        "initial_prompt": "",
        "hotwords": "",
    },
    "diarization": {
        "num_speakers": -1,
        "cluster_threshold": 0.55,
        "window_shift_ratio": 0.1,
        "min_duration_on": 0.3,
        "min_duration_off": 0.5,
    },
    "postprocessing": {
        "merge_max_gap_seconds": 0.45,
    },
}


@dataclass(frozen=True)
class SpeakerInterval:
    start: float
    end: float
    speaker_id: int


@dataclass(frozen=True)
class Word:
    start: float
    end: float
    text: str


@dataclass(frozen=True)
class Segment:
    start: float
    end: float
    speaker: str
    text: str


def analyze_call(filename: str) -> Path:
    """Локально проанализировать разговор и сохранить диалог в Excel.

    Args:
        filename: Абсолютный или относительный путь к аудиофайлу. Относительный
            путь ищется в текущей папке и рядом с этим скриптом.

    Returns:
        Абсолютный путь к ``<имя записи> transcript.xlsx``.

    До первого запуска выполните ``python download_models.py``. Основная
    функция использует только файлы из локальной папки ``models``.
    """
    settings = _load_settings()
    audio_path = _resolve_audio_path(filename)
    _enforce_file_size_limit(audio_path, settings["limits"]["max_file_size_mb"])
    _validate_local_models()

    samples = _decode_mono_16khz(audio_path, settings["limits"]["max_duration_minutes"])
    speaker_intervals = _diarize(samples, settings["diarization"])
    words = _transcribe(audio_path, settings["recognition"])
    dialogue = _combine_words_and_speakers(
        words,
        speaker_intervals,
        settings["postprocessing"]["merge_max_gap_seconds"],
    )

    output_path = audio_path.with_name(f"{audio_path.stem} transcript.xlsx")
    _write_excel(dialogue, audio_path, output_path)
    return output_path


def _load_settings() -> dict[str, Any]:
    _ensure_settings_file()
    settings = deepcopy(DEFAULT_SETTINGS)
    try:
        custom_settings = json.loads(SETTINGS_FILE.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Ошибка JSON в {SETTINGS_FILE.name}, строка {exc.lineno}, столбец {exc.colno}: {exc.msg}"
        ) from exc
    if not isinstance(custom_settings, dict):
        raise ValueError(f"Корень файла {SETTINGS_FILE.name} должен быть JSON-объектом.")
    _merge_known_settings(settings, custom_settings)
    _validate_settings(settings)
    return settings


def _ensure_settings_file() -> None:
    if SETTINGS_FILE.exists():
        if not SETTINGS_FILE.is_file():
            raise ValueError(f"Путь {SETTINGS_FILE} должен указывать на JSON-файл.")
        return

    try:
        with SETTINGS_FILE.open("x", encoding="utf-8", newline="\n") as settings_file:
            json.dump(DEFAULT_SETTINGS, settings_file, ensure_ascii=False, indent=2)
            settings_file.write("\n")
    except FileExistsError:
        # Another process may have created the settings file at the same time.
        pass
    except OSError as exc:
        raise OSError(
            f"Не удалось создать {SETTINGS_FILE.name} рядом со скриптом: {exc}"
        ) from exc


def _merge_known_settings(target: dict[str, Any], source: dict[str, Any], prefix: str = "") -> None:
    for key, value in source.items():
        full_name = f"{prefix}.{key}" if prefix else key
        if key not in target:
            raise ValueError(f"Неизвестная настройка: {full_name}")
        if isinstance(target[key], dict):
            if not isinstance(value, dict):
                raise ValueError(f"Настройка {full_name} должна быть JSON-объектом.")
            _merge_known_settings(target[key], value, full_name)
        else:
            target[key] = value


def _validate_settings(settings: dict[str, Any]) -> None:
    limits = settings["limits"]
    recognition = settings["recognition"]
    diarization = settings["diarization"]
    postprocessing = settings["postprocessing"]

    _require_number(limits["max_file_size_mb"], "limits.max_file_size_mb", minimum=0)
    _require_number(limits["max_duration_minutes"], "limits.max_duration_minutes", minimum=0)

    language = recognition["language"]
    if language is not None and (not isinstance(language, str) or not language.strip()):
        raise ValueError("recognition.language должен быть непустой строкой или null.")
    if recognition["device"] not in {"cpu", "cuda", "auto"}:
        raise ValueError("recognition.device должен быть cpu, cuda или auto.")
    if recognition["compute_type"] not in {"int8", "int8_float16", "float16", "float32", "default"}:
        raise ValueError(
            "recognition.compute_type должен быть int8, int8_float16, float16, float32 или default."
        )
    _require_integer(recognition["cpu_threads"], "recognition.cpu_threads", minimum=0, maximum=128)
    _require_integer(recognition["beam_size"], "recognition.beam_size", minimum=1, maximum=20)
    _require_number(recognition["temperature"], "recognition.temperature", minimum=0, maximum=1)
    _require_boolean(recognition["condition_on_previous_text"], "recognition.condition_on_previous_text")
    _require_boolean(recognition["vad_filter"], "recognition.vad_filter")
    _require_integer(
        recognition["vad_min_silence_duration_ms"],
        "recognition.vad_min_silence_duration_ms",
        minimum=0,
        maximum=10_000,
    )
    for key in ("initial_prompt", "hotwords"):
        if not isinstance(recognition[key], str):
            raise ValueError(f"recognition.{key} должен быть строкой.")

    num_speakers = diarization["num_speakers"]
    _require_integer(num_speakers, "diarization.num_speakers", minimum=-1, maximum=20)
    if num_speakers == 0:
        raise ValueError("diarization.num_speakers должен быть -1 или числом от 1 до 20.")
    _require_number(diarization["cluster_threshold"], "diarization.cluster_threshold", minimum=0, maximum=1)
    _require_number(
        diarization["window_shift_ratio"],
        "diarization.window_shift_ratio",
        minimum=0.01,
        maximum=1,
    )
    _require_number(diarization["min_duration_on"], "diarization.min_duration_on", minimum=0, maximum=10)
    _require_number(diarization["min_duration_off"], "diarization.min_duration_off", minimum=0, maximum=10)
    _require_number(
        postprocessing["merge_max_gap_seconds"],
        "postprocessing.merge_max_gap_seconds",
        minimum=0,
        maximum=30,
    )


def _require_number(value: Any, name: str, minimum: float, maximum: float | None = None) -> None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{name} должен быть числом.")
    if value < minimum or (maximum is not None and value > maximum):
        upper = f" до {maximum}" if maximum is not None else ""
        raise ValueError(f"{name} должен быть в диапазоне от {minimum}{upper}.")


def _require_integer(value: Any, name: str, minimum: int, maximum: int) -> None:
    if isinstance(value, bool) or not isinstance(value, int):
        raise ValueError(f"{name} должен быть целым числом.")
    if not minimum <= value <= maximum:
        raise ValueError(f"{name} должен быть в диапазоне от {minimum} до {maximum}.")


def _require_boolean(value: Any, name: str) -> None:
    if not isinstance(value, bool):
        raise ValueError(f"{name} должен быть true или false.")


def _enforce_file_size_limit(audio_path: Path, max_file_size_mb: float) -> None:
    if max_file_size_mb <= 0:
        return
    actual_size_mb = audio_path.stat().st_size / (1024 * 1024)
    if actual_size_mb > max_file_size_mb:
        raise ValueError(
            f"Размер аудиофайла {actual_size_mb:.1f} МБ превышает лимит "
            f"limits.max_file_size_mb={max_file_size_mb}."
        )


def _resolve_audio_path(filename: str) -> Path:
    if not isinstance(filename, str) or not filename.strip():
        raise ValueError("Имя аудиофайла должно быть непустой строкой.")
    requested = Path(filename).expanduser()
    candidates = [requested]
    if not requested.is_absolute():
        candidates.append(SCRIPT_DIR / requested)
    audio_path = next((item.resolve() for item in candidates if item.is_file()), None)
    if audio_path is None:
        raise FileNotFoundError(f"Аудиофайл не найден: {filename}")
    if audio_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        formats = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Неподдерживаемый формат {audio_path.suffix!r}. Допустимы: {formats}.")
    return audio_path


def _validate_local_models() -> None:
    missing = [
        str(path.relative_to(SCRIPT_DIR))
        for path, expected_size in EXPECTED_MODEL_SIZES.items()
        if not path.is_file() or path.stat().st_size != expected_size
    ]
    if missing:
        details = "\n  - ".join(missing)
        raise RuntimeError(
            "Локальные модели не установлены. Выполните один раз: "
            "python download_models.py\nНе найдены:\n  - " + details
        )


def _decode_mono_16khz(audio_path: Path, max_duration_minutes: float = 0):
    """Декодировать любой поддерживаемый PyAV формат в mono float32/16 kHz."""
    try:
        import av
        import numpy as np
    except ImportError as exc:
        raise RuntimeError("Не установлены зависимости. Выполните: pip install -r requirements.txt") from exc

    chunks = []
    decoded_samples = 0
    max_samples = int(max_duration_minutes * 60 * SAMPLE_RATE) if max_duration_minutes > 0 else 0
    resampler = av.audio.resampler.AudioResampler(format="fltp", layout="mono", rate=SAMPLE_RATE)
    try:
        with av.open(str(audio_path)) as container:
            if not container.streams.audio:
                raise ValueError(f"В файле нет аудиодорожки: {audio_path}")
            for frame in container.decode(container.streams.audio[0]):
                converted = resampler.resample(frame)
                converted_frames = converted if isinstance(converted, list) else [converted]
                for converted_frame in converted_frames:
                    if converted_frame is not None:
                        chunk = converted_frame.to_ndarray().reshape(-1)
                        decoded_samples += chunk.size
                        if max_samples and decoded_samples > max_samples:
                            raise ValueError(
                                f"Длительность аудио превышает лимит "
                                f"limits.max_duration_minutes={max_duration_minutes}."
                            )
                        chunks.append(chunk)
            tail = resampler.resample(None)
            tail_frames = tail if isinstance(tail, list) else [tail]
            for converted_frame in tail_frames:
                if converted_frame is not None:
                    chunk = converted_frame.to_ndarray().reshape(-1)
                    decoded_samples += chunk.size
                    if max_samples and decoded_samples > max_samples:
                        raise ValueError(
                            f"Длительность аудио превышает лимит "
                            f"limits.max_duration_minutes={max_duration_minutes}."
                        )
                    chunks.append(chunk)
    except Exception as exc:
        raise ValueError(f"Не удалось прочитать аудиофайл {audio_path.name}: {exc}") from exc

    if not chunks:
        raise ValueError("Аудиофайл не содержит декодируемого звука.")
    return np.ascontiguousarray(np.concatenate(chunks), dtype="float32")


def _diarize(samples, settings: dict[str, Any]) -> list[SpeakerInterval]:
    try:
        import sherpa_onnx
    except ImportError as exc:
        raise RuntimeError("Не установлен sherpa-onnx. Выполните: pip install -r requirements.txt") from exc

    forced_speakers = _optional_speaker_count(settings["num_speakers"])
    config = sherpa_onnx.OfflineSpeakerDiarizationConfig(
        segmentation=sherpa_onnx.OfflineSpeakerSegmentationModelConfig(
            pyannote=sherpa_onnx.OfflineSpeakerSegmentationPyannoteModelConfig(
                model=str(SEGMENTATION_MODEL),
                window_shift_ratio=settings["window_shift_ratio"],
            )
        ),
        embedding=sherpa_onnx.SpeakerEmbeddingExtractorConfig(model=str(EMBEDDING_MODEL)),
        clustering=sherpa_onnx.FastClusteringConfig(
            num_clusters=forced_speakers,
            threshold=settings["cluster_threshold"],
        ),
        min_duration_on=settings["min_duration_on"],
        min_duration_off=settings["min_duration_off"],
    )
    if not config.validate():
        raise RuntimeError("Не удалось инициализировать локальную модель диаризации.")

    diarizer = sherpa_onnx.OfflineSpeakerDiarization(config)
    raw_result = diarizer.process(samples).sort_by_start_time()
    result = [
        SpeakerInterval(float(item.start), float(item.end), int(item.speaker))
        for item in raw_result
        if float(item.end) > float(item.start)
    ]
    if not result:
        raise ValueError("Модель не обнаружила речь в аудиофайле.")
    return result


def _optional_speaker_count(configured_value: int) -> int:
    raw_value = os.environ.get("VOICE_NUM_SPEAKERS", "").strip()
    if not raw_value:
        return configured_value
    try:
        value = int(raw_value)
    except ValueError as exc:
        raise ValueError("VOICE_NUM_SPEAKERS должен быть целым числом не меньше 1.") from exc
    if value < 1:
        raise ValueError("VOICE_NUM_SPEAKERS должен быть целым числом не меньше 1.")
    return value


def _transcribe(audio_path: Path, settings: dict[str, Any]) -> list[Word]:
    try:
        from faster_whisper import WhisperModel
    except ImportError as exc:
        raise RuntimeError("Не установлен faster-whisper. Выполните: pip install -r requirements.txt") from exc

    cpu_threads = settings["cpu_threads"] or max(1, min(os.cpu_count() or 4, 8))
    model = WhisperModel(
        str(WHISPER_MODEL_DIR),
        device=settings["device"],
        compute_type=settings["compute_type"],
        cpu_threads=cpu_threads,
        local_files_only=True,
    )
    raw_segments, _ = model.transcribe(
        str(audio_path),
        language=settings["language"],
        beam_size=settings["beam_size"],
        temperature=settings["temperature"],
        word_timestamps=True,
        vad_filter=settings["vad_filter"],
        vad_parameters={"min_silence_duration_ms": settings["vad_min_silence_duration_ms"]},
        condition_on_previous_text=settings["condition_on_previous_text"],
        initial_prompt=settings["initial_prompt"] or None,
        hotwords=settings["hotwords"] or None,
    )

    words: list[Word] = []
    for segment in raw_segments:
        segment_words = list(segment.words or [])
        if segment_words:
            for word in segment_words:
                if word.start is None or word.end is None or not word.word.strip():
                    continue
                words.append(Word(float(word.start), float(word.end), word.word))
        elif segment.text.strip():
            words.append(Word(float(segment.start), float(segment.end), segment.text))
    if not words:
        raise ValueError("Whisper не обнаружил распознаваемую русскую речь.")
    return words


def _combine_words_and_speakers(
    words: list[Word],
    speaker_intervals: list[SpeakerInterval],
    merge_max_gap_seconds: float = 0.45,
) -> list[Segment]:
    speaker_labels: dict[int, str] = {}
    combined: list[Segment] = []

    for word in words:
        speaker_id = _speaker_for_word(word, speaker_intervals)
        if speaker_id not in speaker_labels:
            speaker_labels[speaker_id] = f"Собеседник {_speaker_suffix(len(speaker_labels))}"
        speaker = speaker_labels[speaker_id]

        if (
            combined
            and combined[-1].speaker == speaker
            and word.start - combined[-1].end <= merge_max_gap_seconds
        ):
            previous = combined[-1]
            combined[-1] = Segment(
                previous.start,
                max(previous.end, word.end),
                previous.speaker,
                _append_word(previous.text, word.text),
            )
        else:
            combined.append(Segment(word.start, word.end, speaker, word.text.strip()))

    return [segment for segment in combined if segment.text]


def _speaker_for_word(word: Word, intervals: list[SpeakerInterval]) -> int:
    overlaps: list[tuple[float, float, float, int]] = []
    midpoint = (word.start + word.end) / 2
    word_duration = max(word.end - word.start, 1e-6)
    for interval in intervals:
        overlap = max(0.0, min(word.end, interval.end) - max(word.start, interval.start))
        if overlap > 0:
            interval_duration = max(interval.end - interval.start, 1e-6)
            normalized_overlap = overlap / min(word_duration, interval_duration)
            interval_midpoint = (interval.start + interval.end) / 2
            overlaps.append(
                (
                    normalized_overlap,
                    overlap,
                    -abs(midpoint - interval_midpoint),
                    interval.speaker_id,
                )
            )
    if overlaps:
        return max(overlaps)[3]
    return min(
        intervals,
        key=lambda interval: min(abs(midpoint - interval.start), abs(midpoint - interval.end)),
    ).speaker_id


def _append_word(current: str, addition: str) -> str:
    if not current:
        return addition.strip()
    if addition[:1].isspace():
        return current + addition
    if re.match(r"^[,.;:!?…%)\]}]", addition):
        return current + addition
    return current + " " + addition


def _speaker_suffix(index: int) -> str:
    base = len(CYRILLIC_LABELS)
    result = ""
    number = index
    while True:
        number, remainder = divmod(number, base)
        result = CYRILLIC_LABELS[remainder] + result
        if number == 0:
            return result
        number -= 1


def _format_timestamp(seconds: float) -> str:
    total_milliseconds = max(0, round(seconds * 1000))
    hours, remainder = divmod(total_milliseconds, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    whole_seconds, milliseconds = divmod(remainder, 1000)
    hours_prefix = f"{hours:02d}:" if hours else ""
    return f"{hours_prefix}{minutes:02d}:{whole_seconds:02d}.{milliseconds:03d}"


def _write_excel(segments: Iterable[Segment], audio_path: Path, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Не установлен openpyxl. Выполните: pip install -r requirements.txt") from exc

    rows = list(segments)
    if not rows:
        raise ValueError("Диалог пуст — Excel-файл не создан.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Диалог"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(["Таймлайн", "Собеседник", "Что сказано"])

    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="17365D"))
    sheet.row_dimensions[1].height = 26

    color_by_speaker: dict[str, str] = {}
    for row_number, segment in enumerate(rows, start=2):
        timeline = f"{_format_timestamp(segment.start)}–{_format_timestamp(segment.end)}"
        sheet.append([timeline, segment.speaker, segment.text])
        if segment.speaker not in color_by_speaker:
            color_by_speaker[segment.speaker] = SPEAKER_COLORS[len(color_by_speaker) % len(SPEAKER_COLORS)]
        row_fill = PatternFill("solid", fgColor=color_by_speaker[segment.speaker])
        for cell in sheet[row_number]:
            cell.fill = row_fill
            cell.font = Font(name="Aptos", size=10, color="172B4D")
            cell.border = Border(bottom=thin_gray)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 3)
        sheet.cell(row_number, 1).alignment = Alignment(horizontal="center", vertical="top")
        sheet.cell(row_number, 2).font = Font(name="Aptos", size=10, bold=True, color="172B4D")
        sheet.row_dimensions[row_number].height = max(22, min(75, 18 * (1 + len(segment.text) // 90)))

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 19
    sheet.column_dimensions["C"].width = 80
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    table = Table(displayName="CallTranscript", ref=f"A1:C{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    sheet.add_table(table)
    sheet.oddHeader.center.text = f"Диалог: {audio_path.name}"
    sheet.oddFooter.center.text = "Страница &P из &N"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent, delete=False
        ) as temporary_file:
            temporary_name = temporary_file.name
        workbook.save(temporary_name)
        os.replace(temporary_name, output_path)
    finally:
        workbook.close()
        if temporary_name and os.path.exists(temporary_name):
            os.unlink(temporary_name)


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Локально расшифровать разговор в Excel.")
    parser.add_argument("audio_file", help="Путь к аудиозаписи")
    return parser


def main() -> int:
    args = _build_parser().parse_args()
    try:
        result = analyze_call(args.audio_file)
    except Exception as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1
    print(f"Excel-файл создан: {result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
